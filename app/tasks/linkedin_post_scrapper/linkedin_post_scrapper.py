"""
LinkedIn Login Automation with Selenium

Requirements:
1. Python 3.7+
2. Selenium package: `pip install selenium`
3. ChromeDriver installed and path configured
4. Environment variables set:
   - LINKEDIN_USERNAME: Your LinkedIn email/username
   - LINKEDIN_PASSWORD: Your LinkedIn password

Setup:
1. Download ChromeDriver from https://chromedriver.chromium.org/
2. Add ChromeDriver to PATH or specify path in config
"""

import os
import logging
import re
import sqlite3
import time
from typing import Generator, List, Dict, Optional
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    WebDriverException,
    InvalidCookieDomainException,
)
from dotenv import load_dotenv
import undetected_chromedriver as uc
from selenium_stealth import stealth
import logging
from selenium.common.exceptions import WebDriverException


import pandas as pd
import pickle
from datetime import datetime, timedelta
from selenium.webdriver.common.keys import Keys

from app.tasks.db_utils import check_post_exists, insert_linkedin_post
from app.tasks.linkedin_post_scrapper.utils import FixedSizeStore, LinkedInQueryBuilder, contains_email, read_queries_from_file

load_dotenv(override=True)


# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

# Configuration constants
LINKEDIN_LOGIN_URL = "https://www.linkedin.com/login"
LINKEDIN_FEED_URL = "https://www.linkedin.com/feed/"
# DRIVER_PATH = os.environ.get("CHROME_DRIVER_PATH", "default/path")
# logging.info("Using ChromeDriver path: %s", DRIVER_PATH)

CONFIG = {
    # "driver_path": DRIVER_PATH,
    "headless": True  # Set to True for headless mode
}

# Add these constants to the existing configuration
QUERIES_FILE = "search_queries.txt"
MAX_POST_AGE_WEEKS = 10
OUTPUT_EXCEL_FILE = "linkedin_posts.xlsx"

POST_SELECTOR = "li.artdeco-card"  # More stable class selector
CONTENT_SELECTOR = "[data-test-id='post-content']"


def setup_driver(config):
    """Initialize and configure Chrome WebDriver"""
    try:
        options = webdriver.ChromeOptions()
        if config["headless"]:
            options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
        driver = webdriver.Chrome()
        driver.maximize_window()
        return driver
    except WebDriverException as e:
        logging.error("WebDriver initialization failed: %s", e)
        raise


def login_to_linkedin(driver, username=None, password=None):
    """Handle LinkedIn login with cookies fallback to credentials"""
    try:
        # Try to use existing cookies first
        try:
            driver.get(LINKEDIN_LOGIN_URL)
            cookies = pickle.load(open("cookies.pkl", "rb"))
            for cookie in cookies:
                try:
                    driver.add_cookie(cookie)
                except InvalidCookieDomainException:
                    pass
            logging.info("Cookies loaded successfully")
        except (FileNotFoundError, Exception) as e:
            logging.warning("No cookies found or invalid cookies: %s", str(e))

        # Verify login status
        driver.get(LINKEDIN_FEED_URL)
        time.sleep(2)  # Small wait for potential redirects

        if not is_logged_in(driver):
            logging.info("Session expired or no valid cookies. Starting fresh login...")
            perform_credentials_login(driver, username, password)

            # Verify successful credential login
            if not is_logged_in(driver):
                raise RuntimeError("Credentials login failed")

            # Save new cookies only after successful login
            pickle.dump(driver.get_cookies(), open("cookies.pkl", "wb"))
            logging.info("New cookies saved")

        logging.info("Successfully logged in")
        return True

    except Exception as e:
        logging.error("Login failed: %s", str(e))
        raise


def is_logged_in(driver):
    """Check if we're actually logged in"""
    try:
        # Check for either feed content or login elements
        WebDriverWait(driver, 15).until(
            lambda d: d.find_elements(
                By.CSS_SELECTOR, ".scaffold-finite-scroll__content"
            )
            or d.find_elements(By.ID, "username")
        )
        return bool(
            driver.find_elements(By.CSS_SELECTOR, ".scaffold-finite-scroll__content")
        )
    except TimeoutException:
        return False


def perform_credentials_login(driver, username, password):
    """Handle manual credential login"""
    try:
        driver.get(LINKEDIN_LOGIN_URL)
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "username"))
        )

        # Input credentials
        username_field = driver.find_element(By.ID, "username")
        password_field = driver.find_element(By.ID, "password")

        username_field.clear()
        username_field.send_keys(username)
        password_field.clear()
        password_field.send_keys(password)

        # Submit form
        driver.find_element(By.XPATH, "//button[@type='submit']").click()

        # Handle potential security checks
        WebDriverWait(driver, 30).until(
            lambda d: "checkpoint/challenge" not in d.current_url
        )

    except Exception as e:
        logging.error("Credential login failed: %s", str(e))
        raise


def handle_2fa(driver):
    """Check for 2FA prompt and log occurrence"""
    try:
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "input__phone_verification_pin"))
        )
        logging.warning("2FA authentication required")
        return True
    except (TimeoutException, NoSuchElementException):
        return False


def handle_captcha(driver):
    """Check for CAPTCHA challenge"""
    try:
        driver.find_element(By.ID, "captcha-internal")
        logging.warning("CAPTCHA challenge detected")
        return True
    except NoSuchElementException:
        return False


def check_login_errors(driver):
    """Check for login error messages"""
    try:
        error_message = driver.find_element(
            By.XPATH, "//div[contains(@class, 'alert-error')]"
        )
        if error_message:
            logging.error("Login failed: %s", error_message.text)
            return True
    except NoSuchElementException:
        return False


def dismiss_popups(driver):
    """Handle unexpected pop-ups/modals"""
    try:
        # Example for "Save Password" pop-up
        WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(text(), 'Later')]")
            )
        ).click()
        logging.info("Dismissed unexpected pop-up")
    except TimeoutException:
        pass


def scroll_to_element(driver, element):
    """Enhanced scroll with visibility check"""
    driver.execute_script(
        """
		const element = arguments[0];
		element.scrollIntoView({
			behavior: 'smooth',
			block: 'center',
			inline: 'center'
		});
		
		// Add visual marker for debugging
		element.style.outline = '2px solid #ff0000';
	""",
        element,
    )

    WebDriverWait(driver, 5).until(
        lambda d: element.is_displayed()
        and element.location_once_scrolled_into_view["y"]
        < d.execute_script("return window.innerHeight")
    )


def perform_search(driver, query):
    """Execute LinkedIn search for a given query and navigate to posts"""
    try:
        # Wait for search box and execute search
        search_box = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located(
                (By.XPATH, "//input[contains(@aria-label, 'Search')]")
            )
        )
        search_box.clear()
        search_box.send_keys(query)
        search_box.send_keys(Keys.RETURN)
        logging.info(f"Executed search for: {query}")

        # Navigate to Posts tab
        WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'Posts')]"))
        ).click()
        logging.info("Navigated to Posts section")

        # Wait for post results to load
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "div.feed-shared-update-v2")
            )
        )

    except TimeoutException as e:
        logging.error(f"Search failed for {query}: {str(e)}")
        raise


def parse_timestamp(time_str: str) -> datetime:
    """Convert LinkedIn relative time to datetime object"""
    now = datetime.now()
    try:
        time_str = time_str.lower()
        numbers = re.findall(r"\d+", time_str)
        if not numbers:
            return now

        value = int(numbers[0])

        if "minute" in time_str:
            return now - timedelta(minutes=value)
        if "hour" in time_str:
            return now - timedelta(hours=value)
        if "day" in time_str:
            return now - timedelta(days=value)
        if "week" in time_str:
            return now - timedelta(weeks=value)
        if "month" in time_str:
            return now - timedelta(days=value * 30)
        if "year" in time_str:
            return now - timedelta(days=value * 365)

        logging.info(f"time_str: {time_str} and now: {now}")
        return now
    except Exception as e:
        print(f"Error parsing timestamp: {str(e)}")
        return now


def is_recent_post(post_date, max_weeks=MAX_POST_AGE_WEEKS):
    """Check if post is within date threshold"""
    cutoff_date = datetime.now() - timedelta(weeks=max_weeks)
    return post_date >= cutoff_date


def scroll_to_load_posts(driver):
    """Handle infinite scroll loading"""
    last_height = driver.execute_script("return document.body.scrollHeight")
    scroll_attempts = 0

    while scroll_attempts < 5:  # Limit max scroll attempts
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)  # Allow time for loading

        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height
        scroll_attempts += 1


# def scrape_posts(driver):
# 	"""Extract and filter posts from current page"""
# 	posts_data = []
# 	cutoff_date = datetime.now() - timedelta(weeks=MAX_POST_AGE_WEEKS)

# 	try:
# 		posts = driver.find_elements(By.XPATH, "//div[contains(@class, 'feed-shared-update-v2')]")
# 		logging.info(f"Found {len(posts)} posts")

# 		for post in posts:
# 			try:
# 				# Extract timestamp
# 				time_element = post.find_element(By.XPATH, ".//span[@class='visually-hidden']")
# 				post_date = parse_timestamp(time_element.text)

# 				if not is_recent_post(post_date):
# 					continue

# 				# Extract metadata
# 				content_element = post.find_element(By.XPATH, ".//div[contains(@class, 'feed-shared-update-v2__description')]")
# 				content = content_element.text
# 				author_element = post.find_element(By.XPATH, ".//span[contains(@class, 'update-components-actor__title')]/a")
# 				author = author_element.text
# 				profile_link_element = post.find_element(By.XPATH, ".//a[contains(@class, 'update-components-actor__container-link')]")
# 				profile_link = profile_link_element.get_attribute('href')

# 				# Get engagement metrics
# 				try:
# 					reactions_element = post.find_element(By.XPATH, ".//span[contains(@class, 'social-details-social-counts__reactions')]")
# 					reactions = reactions_element.text
# 				except NoSuchElementException:
# 					reactions = '0'

# 				posts_data.append({
# 					'query': driver.current_query,
# 					'content': content,
# 					'timestamp': post_date.strftime('%Y-%m-%d %H:%M'),
# 					'author': author,
# 					'profile_link': profile_link,
# 					'reactions': reactions
# 				})

# 			except NoSuchElementException as e:
# 				logging.warning(f"Skipping post due to missing elements: {str(e)}")
# 				continue

# 	except NoSuchElementException as e:
# 		logging.error(f"No posts found: {str(e)}")

# 	return posts_data


def save_to_excel(data, filename="output.xlsx", sheet_name="Sheet1"):
    """Save scraped data to Excel file with safe and unique sheet name"""
    if not data:
        logging.warning("No data to save")
        return

    # Ensure sheet name is max 31 characters
    base_name = sheet_name[:31]

    # Check if sheet exists and generate a unique one if needed
    def get_unique_sheet_name(file, base):
        if not os.path.exists(file):
            return base
        try:
            wb = load_workbook(file)
            if base not in wb.sheetnames:
                return base
            i = 1
            while True:
                suffix = f"_{i}"
                candidate = base[:31 - len(suffix)] + suffix
                if candidate not in wb.sheetnames:
                    return candidate
                i += 1
        except Exception as e:
            logging.error(f"Failed to load workbook: {e}")
            return base  # fallback

    unique_sheet_name = get_unique_sheet_name(filename, base_name)

    df = pd.DataFrame(data)
    if not os.path.exists(filename):
        df.to_excel(filename, index=False, sheet_name=unique_sheet_name)
    else:
        with pd.ExcelWriter(filename, mode='a', engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=unique_sheet_name)

    logging.info(f"Data saved to {filename} in sheet '{unique_sheet_name}'")

def scrape_linkedin_post(html):
    soup = BeautifulSoup(html, "html.parser")

    data = {
        "author_name": None,
        "profile_url": None,
        "post_date": datetime.now(),  # Default to current time
        "post_link": None,
        "post_text": None,
        "hashtags": [],
        "reactions": None,
        "comments": None,
        "query": None
    }

    try:
        # Author Section
        author_section = soup.find(
            "div", class_=lambda x: x and "update-components-actor" in x
        )
        if author_section:
            # Extract Author Name
            name_tag = author_section.select_one(
                'span[dir="ltr"] span[aria-hidden="true"]'
            )
            if name_tag:
                data["author_name"] = name_tag.get_text(strip=True)

            # Extract Profile URL
            profile_link = author_section.find(
                "a", class_=lambda x: x and "update-components-actor__meta-link" in x
            )
            if profile_link:
                data["profile_url"] = profile_link.get("href")

            # Extract Post Date
            date_container = author_section.find(
                "span", class_="update-components-actor__sub-description"
            )
            if date_container:
                hidden_date = date_container.find("span", class_="visually-hidden")
                if hidden_date:
                    relative_time = hidden_date.get_text(strip=True)
                    logging.debug(f"Relative time: {relative_time}")
                    data["post_date"] = parse_timestamp(relative_time)

        # Post Link: First try to find a link wrapper.
        link_wrapper = soup.find(
            "div", class_="update-components-update-v2__link-wrapper"
        )
        if link_wrapper:
            logging.debug("Found link wrapper: " + link_wrapper.prettify())
            post_link_tag = link_wrapper.find("a", href=True)
            if post_link_tag:
                data["post_link"] = post_link_tag.get("href")
        else:
            logging.debug("Link wrapper not found using primary selector")

        # Fallback: Use the data-urn from the feed-shared-update-v2 container.
        if not data["post_link"]:
            post_container = soup.find(
                "div", class_=lambda x: x and "feed-shared-update-v2" in x
            )
            if post_container and post_container.has_attr("data-urn"):
                urn = post_container["data-urn"]
                # Construct the post link using the URN.
                data["post_link"] = f"https://www.linkedin.com/feed/update/{urn}"
                logging.debug(
                    "Using data-urn fallback for post link: " + data["post_link"]
                )

        # Post Text
        text_container = soup.find("div", class_="update-components-text")
        if text_container:
            data["post_text"] = " ".join(text_container.stripped_strings)
            data["post_text"] = (
                data["post_text"].encode("utf-8", errors="replace").decode("utf-8")
            )
        else:
            logging.warning(f"No text container found for post: {data}")
            with open("no_text_container_found.txt", "a") as f:
                f.write(html)
                f.write("===============\n\n\n")
            # input("Written to file Press Enter to continue...")
            return None

        # Hashtags
        data["hashtags"] = [
            tag.get_text(strip=True)
            for tag in soup.select('a[href*="/hashtag/"]')
            if tag.text
        ]

        # Engagement Metrics: Reactions and Comments
        reactions = soup.find(
            "button", attrs={"aria-label": re.compile(r"reactions", re.I)}
        )
        if reactions:
            data["reactions"] = reactions.get_text(strip=True)
        comments = soup.find(
            "button", attrs={"aria-label": re.compile(r"comments", re.I)}
        )
        if comments:
            # We assume the first token is the comment count.
            data["comments"] = comments.get_text(strip=True).split()[0]

    except Exception as e:
        logging.error(f"Parsing error: {str(e)}")

    return data
    
def scrape_linkedin_posts_for_query(driver, query: str, max_results: int, sort_by_latest: bool = True):
    """
    Main scraping function that:
    1. Loads all posts through infinite scroll
    2. Processes posts after loading completes
    3. Stops when either 300 posts collected or oldest post >10 weeks old
    """
    cutoff_date = datetime.now() - timedelta(weeks=10)
    total_scrapped_posts = 0

    def sort_by_latest_function(driver):
        """
        Sort search results by most recent posts.

        This function performs the following steps:
        1. Clicks the "Sort by" button using its unique ID.
        2. Waits for the dropdown to appear and clicks the "Latest" option
                (identified via the label with for='sortBy-date_posted').
        3. Clicks the "Show results" button to apply the filter.

        JavaScript clicks are used along with scrolling to ensure the elements
        are in view and interactable.
        """
        try:
            # Step 1: Click the "Sort by" button
            sort_button = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.ID, "searchFilter_sortBy"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", sort_button)
            driver.execute_script("arguments[0].click();", sort_button)
            logging.info("Clicked on 'Sort by' button")

            # Step 2: Wait for the "Latest" option and click it
            # The "Latest" option is represented by the label for the radio input with id "sortBy-date_posted"
            latest_option = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//label[@for='sortBy-date_posted']")
                )
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", latest_option)
            driver.execute_script("arguments[0].click();", latest_option)
            logging.info("Selected 'Latest' sort option")

            # Step 3: Click the "Show results" button to apply the filter
            show_results_button = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable(
                    (
                        By.XPATH,
                        "//button[@aria-label='Apply current filter to show results']",
                    )
                )
            )
            driver.execute_script(
                "arguments[0].scrollIntoView(true);", show_results_button
            )
            driver.execute_script("arguments[0].click();", show_results_button)
            logging.info("Clicked on 'Show results' button")

            time.sleep(2)  # Allow time for the sort filter to be applied
        except Exception as e:
            logging.error(f"Failed to sort by latest: {str(e)}")
            raise

    def scroll_to_bottom():
        """Scroll to page bottom and wait for loading"""
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2.5)

    def get_visible_posts():
        """Get currently visible post elements"""
        return driver.find_elements(
            By.XPATH, "//li[contains(@class, 'artdeco-card mb2')]"
        )

    def get_post_age(post_element):
        """Extract post age from element return timedelta"""
        try:
            time_element = post_element.find_element(
                By.XPATH,
                ".//span[contains(@class, 'feed-shared-actor__sub-description')]//span[@aria-hidden]",
            )
            return parse_timestamp(time_element.text)
        except NoSuchElementException:
            return datetime.now()

    def should_stop_loading(posts):
        """Check loading stop conditions"""
        if len(posts) >= 1:
            logging.info(f"Reached {max_results} posts during loading")
            return True
        if posts:
            try:
                oldest_age = get_post_age(posts[-1])
                if oldest_age < cutoff_date:
                    logging.info("Found post older than 10 weeks during loading")
                    return True
            except Exception as e:
                logging.warning(f"Age check error: {str(e)}")
        return False

    scroll_attempts = 0
    max_attempts = 15
    previous_post_count = 0
    recent_posts_store = FixedSizeStore(10)

    
    try:
        perform_search(driver, query)
        if sort_by_latest:
            sort_by_latest_function(driver)
        time.sleep(2)

        while scroll_attempts < max_attempts and total_scrapped_posts < max_results:
            # Get current batch of posts and count
            current_posts = get_visible_posts()
            current_count = len(current_posts)

            # Calculate new posts added since last check
            new_posts_count = current_count - previous_post_count
            logging.info(f"Current post count: {current_count}")
            logging.info(f"New posts detected: {new_posts_count}")

            if new_posts_count > 0:
                # Process only the newly added posts (last N elements)
                new_posts = current_posts[-(new_posts_count+2):]
                keep_processing = True

                with open("new_posts.html", "w", encoding="utf-8") as f:
                    f.write("\n\n\n\n\n".join([post.get_attribute("outerHTML") for post in new_posts]))

                for post in new_posts:
                    print("-" * 50)
                    print("-" * 50)
                    if not keep_processing or total_scrapped_posts >= max_results:
                        logging.info(
                            f"Reached {max_results} posts or stopped processing"
                        )
                        break

                    try:
                        post_html = post.get_attribute("outerHTML")
                        with open("post.html", "w", encoding="utf-8") as f:
                            f.write(post_html)
                        post_data = scrape_linkedin_post(post_html)
                        

                        if (
                            not post_data
                            or not post_data["post_text"]
                            or not contains_email(post_data["post_text"])
                        ):
                            logging.info("No post text or no email found")
                            continue

                        if recent_posts_store.find(post_data["post_link"]):
                            logging.info("Already scrapped this post")
                            continue
                        recent_posts_store.insert(post_data["post_link"])
                        
                        if check_post_exists(post_data["post_link"]):
                            logging.info("Already scrapped this post")
                            continue

                        # with open("already_scrapped.txt", "a") as f:
                        #     f.write(f"{post_data['post_link']}\n")
                        post_date = post_data.get("post_date", datetime.now())
                        post_data["query"] = query

                        # Age check for each new post
                        if post_date < cutoff_date:
                            logging.info("Reached posts older than 10 weeks")
                            keep_processing = False
                            break

                        logging.info(f"Yielding post: {post_data['post_link']}")
                        yield post_data
                        total_scrapped_posts += 1
                        logging.info(f"Collected {total_scrapped_posts} posts")

                    except Exception as e:
                        import traceback
                        traceback.print_exc()
                        logging.warning(f"Post processing error: {str(e)}")
                        continue

                # Update count tracker
                previous_post_count = current_count
                scroll_attempts = 0  # Reset counter on new posts
            else:
                # No new posts - try alternative loading
                scroll_attempts += 1
                logging.info(f"Scroll attempts: {scroll_attempts}/{max_attempts}")

                try:
                    # Attempt click on "Show more results" if available
                    driver.find_element(
                        By.XPATH, "//button[contains(., 'Show more results')]"
                    ).click()
                    time.sleep(2)
                except NoSuchElementException:
                    pass

            # Scroll regardless of new content to trigger loading
            scroll_to_bottom()
            time.sleep(2.5)

    except Exception as e:
        logging.error(f"Scraping interrupted: {str(e)}")

    logging.info(f"Finished with {total_scrapped_posts} posts collected")


def scrape_linkedin_feed(
    username: str,
    password: str,
    queries_file: str,
    config: Optional[dict] = CONFIG,
    save_path: Optional[str] = OUTPUT_EXCEL_FILE,
) -> Generator[List[Dict], None, None]:
    """
    Scrapes LinkedIn posts based on search queries and yields batches of results

    Args:
            username: LinkedIn account username
            password: LinkedIn account password
            queries_file: Path to file containing search queries (one per line)
            config: Browser configuration dictionary (default: DEFAULT_CONFIG)
            save_path: Optional path to save all results in Excel format

    Yields:
            Batches of post dictionaries as they are scraped

    Raises:
            RuntimeError: For authentication failures or critical errors
            ValueError: For missing required parameters
    """
    driver = None
    all_posts = []

    try:
        # Validate credentials
        if not username or not password:
            raise ValueError("Both username and password must be provided")

        # Set default config if not provided
        # Initialize browser driver
        driver = setup_driver(config)
        logging.info("Browser initialized with config: %s", config)

        # Perform login sequence
        login_to_linkedin(driver, username, password)
        time.sleep(2)  # Allow for post-login navigation

        # Check for login-related issues
        if check_login_errors(driver):
            raise RuntimeError("Login failed - invalid credentials detected")

        if handle_2fa(driver):
            raise RuntimeError(
                "2FA authentication required - please authenticate manually"
            )

        if handle_captcha(driver):
            logging.warning("CAPTCHA challenge detected, manual intervention needed")
            time.sleep(15)

        dismiss_popups(driver)
        logging.info("Successfully logged in to LinkedIn")

        # Process search queries
        queries_with_max_results = read_queries_from_file(queries_file)
        logging.info("Loaded %d search queries from %s",
                     len(queries_with_max_results), queries_file)

        for query, max_results, sort_by_latest in queries_with_max_results:
            try:
                logging.info("Processing query: '%s'", query)
                for batch in scrape_linkedin_posts_for_query(driver, query, max_results, sort_by_latest):
                    yield batch
                    all_posts.append(batch)
            except Exception as e:
                logging.error("Failed to process query '%s': %s", query, str(e))
                continue
        # Save all posts to Excel in new sheet with date as 3-jan_query
        sheet_name = query + "_" + datetime.now().strftime("%d_%b_%y")
        save_to_excel(all_posts, save_path, sheet_name)
        logging.info("Saved %d total posts to %s", len(all_posts), save_path)

    except Exception as e:
        logging.error("Scraping process failed: %s", str(e))
        raise RuntimeError(f"Scraping terminated: {str(e)}") from e
    finally:
        if driver:
            driver.quit()
            logging.info("Browser instance closed")


if __name__ == "__main__":

    current_dir = os.path.dirname(__file__)
    QUERIES_FILE_PATH = os.path.join(
        current_dir, "search_queries.yaml")
    print(f"Using queries file: {QUERIES_FILE_PATH}")
    for batch in scrape_linkedin_feed(
        username=os.getenv("LINKEDIN_USERNAME"),
        password=os.getenv("LINKEDIN_PASSWORD"),
        queries_file=QUERIES_FILE_PATH,
    ):
        print(f"Yielded batch of {len(batch)} posts")
        print("-" * 50)
